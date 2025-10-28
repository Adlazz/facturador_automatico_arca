import pandas as pd
from dataclasses import dataclass
from typing import Optional
from datetime import datetime
from openpyxl import load_workbook
import os

@dataclass
class FacturaData:
    """Estructura de datos para cada factura"""
    cliente: str
    cuit: str
    cond_iva: str
    importe: float
    iva: float
    rendicion: str
    fecha: datetime
    periodo: str
    realizado: bool
    row_index: int
    pto_venta: str

class ExcelHandler:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.df = None
        self.workbook = None
        self.sheet = None

    def load_excel(self) -> bool:
        try:
            # Especificar dtype para la columna CUIT
            dtype_dict = {'CUIT': str}
            
            # Cargar Excel con converters específicos para CUIT
            self.df = pd.read_excel(
                self.excel_path,
                header=0,
                usecols="A:J",
                dtype=dtype_dict,
                converters={
                    'CUIT': lambda x: str(x).split('.')[0],
                    'Cliente': lambda x: str(x).strip() if pd.notna(x) else ''
                }
            )
            
            # Limpiar nombres de columnas
            self.df.columns = self.df.columns.str.strip()
            
            # Filtrar filas donde Cliente no está vacío
            self.df = self.df[self.df['Cliente'].astype(str).str.strip() != '']
            
            # Limpiar otros valores nulos con valores por defecto
            self.df = self.df.fillna({
                'CUIT': '',
                'Cond_IVA': '',
                'Importe': 0,
                'IVA': 0,
                'Rendicion': '',
                'Periodo': '',
                'Realizado': '',
                'Pto Venta': '4'
            })
            
            # Cargar workbook
            self.workbook = load_workbook(self.excel_path)
            self.sheet = self.workbook.active
            
            self._validate_required_columns()
            return True
            
        except Exception as e:
            print(f"Error al cargar el archivo Excel: {str(e)}")
            return False

    def _validate_required_columns(self):
        """Valida que existan las columnas necesarias"""
        required_columns = [
            'Cliente', 'CUIT', 'Cond_IVA', 'Importe', 
            'IVA', 'TOTAL', 'Rendicion', 'Fecha', 'Periodo'
        ]
        
        # Mostrar columnas disponibles para diagnóstico
        print("Columnas encontradas:", list(self.df.columns))
        
        missing_columns = [col for col in required_columns if col not in self.df.columns]
        if missing_columns:
            raise ValueError(f"Faltan las siguientes columnas requeridas: {missing_columns}")

    def limpiar_cuit(self, cuit_raw):
        """Limpia y formatea el CUIT correctamente"""
        try:
            if pd.isna(cuit_raw) or str(cuit_raw).strip() == '':
                return ''
                
            # Convertir a string y eliminar espacios
            cuit_str = str(cuit_raw).strip()
            
            # Eliminar cualquier carácter no numérico
            cuit_limpio = ''.join(c for c in cuit_str if c.isdigit())
            
            # Si el CUIT tiene más de 11 dígitos, tomar solo los primeros 11
            if len(cuit_limpio) > 11:
                cuit_limpio = cuit_limpio[:11]
            
            return cuit_limpio
            
        except Exception as e:
            print(f"Error limpiando CUIT {cuit_raw}: {str(e)}")
            return ''

    def get_facturas_pendientes(self) -> list[FacturaData]:
        """Obtiene solo las facturas no realizadas"""
        facturas = []
        
        for index, row in self.df.iterrows():
            # Verificar si el cliente está vacío
            if pd.isna(row['Cliente']) or str(row['Cliente']).strip() == '':
                continue
            
            # Verificar si la factura está realizada
            realizado = pd.notna(row.get('Realizado')) and str(row['Realizado']).strip() in ['✓', 'True']
            
            if not realizado:
                try:
                    # Usar el CUIT ya limpio del DataFrame
                    cuit = self.limpiar_cuit(row['CUIT'])
                    if not cuit:
                        print(f"CUIT vacío para cliente {row['Cliente']}, saltando...")
                        continue
                    
                    # Convertir valores numéricos
                    try:
                        importe = float(str(row['Importe']).replace(',', '.'))
                        iva = float(str(row['IVA']).replace(',', '.'))
                    except ValueError:
                        print(f"Error en valores numéricos para cliente {row['Cliente']}, saltando...")
                        continue

                    # Verificar fecha
                    fecha = pd.to_datetime(row['Fecha'])
                    if pd.isna(fecha):
                        print(f"Fecha inválida para cliente {row['Cliente']}, saltando...")
                        continue

                    factura = FacturaData(
                        cliente=str(row['Cliente']).strip(),
                        cuit=cuit,
                        cond_iva=str(row['Cond_IVA']).strip().upper(),
                        importe=importe,
                        iva=iva,
                        rendicion=str(row['Rendicion']).strip().split('.')[0],
                        fecha=fecha.to_pydatetime(),
                        periodo=str(row['Periodo']).strip(),
                        realizado=False,
                        row_index=index + 2,
                        pto_venta=str(row.get('Pto Venta', '4')).strip().split('.')[0]
                    )
                    
                    # Solo agregar si todos los datos son válidos
                    if self.validate_factura_data(factura):
                        facturas.append(factura)
                        print(f"Factura agregada para {factura.cliente}: CUIT={cuit}, Importe={importe}")
                    else:
                        print(f"Factura no válida para {row['Cliente']}, saltando...")
                        
                except Exception as e:
                    print(f"Error procesando fila {index + 2}: {str(e)}")
                    continue
        
        return facturas

    def validate_factura_data(self, factura: FacturaData) -> bool:
        """Realiza validaciones adicionales sobre los datos de la factura"""
        try:
            # Validar que el cliente no esté vacío
            if not factura.cliente or factura.cliente.strip() == '':
                print("Error: Cliente vacío")
                return False

            # Validación del CUIT
            cuit_limpio = self.limpiar_cuit(factura.cuit)
            if len(cuit_limpio) != 11:
                print(f"CUIT inválido para {factura.cliente}: '{factura.cuit}' (limpio: '{cuit_limpio}')")
                return False
            
            # Asignar el CUIT limpio
            factura.cuit = cuit_limpio
            
            # Validación de condición de IVA
            condiciones_validas = ['RI', 'CF', 'M', 'E']
            if factura.cond_iva not in condiciones_validas:
                print(f"Condición de IVA inválida para {factura.cliente}: '{factura.cond_iva}' - debe ser una de {condiciones_validas}")
                return False
            
            # Validación de importes
            try:
                importe = float(str(factura.importe).replace(',', '.'))
                assert importe > 0, f"Importe debe ser mayor a 0 para {factura.cliente}"
            except (ValueError, AssertionError) as e:
                print(f"Error en importe para {factura.cliente}: {str(e)}")
                return False
            
            # Validación de rendición
            if not factura.rendicion or str(factura.rendicion).strip() == '':
                print(f"Error: Número de rendición vacío para {factura.cliente}")
                return False

            # Validación de fecha
            if factura.fecha is None:
                print(f"Error: Fecha inválida para {factura.cliente}")
                return False

            # Validación de periodo
            if not factura.periodo or str(factura.periodo).strip() == '':
                print(f"Error: Periodo vacío para {factura.cliente}")
                return False
            
            print(f"Validación exitosa para {factura.cliente} - CUIT: {factura.cuit}")
            return True
                
        except Exception as e:
            print(f"Error inesperado validando factura para {factura.cliente}: {str(e)}")
            return False

    def marcar_como_realizada(self, factura: FacturaData) -> bool:
        """Marca una factura como realizada en el Excel"""
        try:
            # Encontrar la columna 'Realizado'
            realizado_col = None
            for idx, cell in enumerate(self.sheet[1], 1):
                if cell.value == 'Realizado':
                    realizado_col = idx
                    break
            
            if realizado_col is None:
                raise ValueError("No se encontró la columna 'Realizado'")
            
            # Marcar como realizada
            self.sheet.cell(row=factura.row_index, column=realizado_col, value='✓')
            
            # Guardar el archivo
            self.workbook.save(self.excel_path)
            print(f"Factura de {factura.cliente} marcada como realizada")
            return True
            
        except Exception as e:
            print(f"Error al marcar factura como realizada: {str(e)}")
            return False

def main():
    """Función principal para pruebas"""
    excel_handler = ExcelHandler("facturador_test.xlsx")
    if excel_handler.load_excel():
        facturas = excel_handler.get_facturas_pendientes()
        print(f"\nSe encontraron {len(facturas)} facturas pendientes")
        
        for factura in facturas:
            print(f"\nProcesando factura para {factura.cliente}...")
            excel_handler.marcar_como_realizada(factura)

if __name__ == "__main__":
    main()